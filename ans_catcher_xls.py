# -*- coding: utf-8 -*-
# 有ㄌ上面這行好像就可以打中文註解

# remember to 'pip install openpyxl'
from openpyxl import load_workbook

wb = load_workbook('未命名表單 (回應).xlsx')
print(wb.sheetnames)


# grab the active worksheet (first)
ws = wb.active

# 座號（要注意是從0開始編號）
Stu_num = 0

# 題號（要注意是從0開始編號）
Que_num = 0

# 這是最後要給沛儒的資料，二維陣列
Ans_list = [[0]*30 for i in range(45)]

# 45列(座號)、30行(題數)
# 引用方法： Ans_list[Stu_num][Que_num]


for row in ws.iter_rows(min_row=2, min_col=7, values_only=True):

    # DEBUG
    print(row)
    print(">> 座號序號：" + str(Stu_num))
    print(">> 題目數量：" + str(len(Ans_list[Stu_num])))

    Ans_list[Stu_num] = list(row)

    Que_num = 0
    for Que_num in range(len(Ans_list[Stu_num])):

        # DEBUG
        print("題目序號：" + str(Que_num))
        print(Ans_list[Stu_num][Que_num])

        if(type(Ans_list[Stu_num][Que_num]) != int and len(Ans_list[Stu_num][Que_num]) > 1):
            # 如果收到的答案不只一個字母，就把所有答案合成連續的字母
            # eg. 收到'A, B, C' 轉成'ABC'

            print(len(Ans_list[Stu_num][Que_num]))

            temp_str = ""

            for n in range(0, len(Ans_list[Stu_num][Que_num]), 3):
                temp_str = temp_str + str(Ans_list[Stu_num][Que_num][n])

            print(temp_str)
            Ans_list[Stu_num][Que_num] = temp_str

    print(Ans_list[Stu_num])
    Stu_num += 1

# 輸出第一筆資料、第一題
print(Ans_list[0][0])
