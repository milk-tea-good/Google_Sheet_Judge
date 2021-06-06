# remember to 'pip install openpyxl'
from openpyxl import load_workbook
from selenium.webdriver.support import expected_conditions

wb = load_workbook('Test005.xlsx')      # 讀入 Test005.xlsx
ws = wb.active  # grab the active worksheet (first)


def mutiple_selection(ans, stu):    # 計算多選題錯的選項數的函式
    # initialization
    option = []
    for a in range(0, 5):
        option.append(1)
    counter = 0

    for i in range(len(ans)):
        temp1 = ans[i:i+1]
        if(60 <= ord(temp1)):
            option[ord(temp1)-65] *= (-1)

    for i in range(len(stu)):
        temp2 = stu[i:i+1]
        if(60 <= ord(temp2)):
            option[ord(temp2)-65] *= (-1)

    for i in range(0, 5):
        if(option[i] != 1):
            counter -= -1

    return counter      # 回傳錯的選項數


Stu_num = 0     # 座號(要注意是從0開始編號)
Que_num = 0     # 題號(要注意是從0開始編號)

num_choice = int(input("單選題總題數："))       # 單選題數
num_selection = int(input("多選題總題數："))    # 多選題數
print("單選：1～%d  多選：%d～%d " %
      (num_choice, num_choice+1, num_choice+num_selection))
total_score = int(input("此試券總分："))        # 總分
scr_choice = int(input("單選單題配分："))       # 單選配分
scr_selection = int(input("多選單題配分："))    # 多選配分
# compare with row[0]
stu_number = int(input("學生人數："))


# 這是最後要給沛儒的資料，二維陣列
Ans_list = [[0]*(num_choice+num_selection) for i in range(stu_number+1)]
# 引用方法： Ans_list[Stu_num][Que_num]

for row in ws.iter_rows(min_row=2, min_col=6, values_only=True):

    try:
        Ans_list[Stu_num] = list(row)
        Que_num = 0
        for Que_num in range(len(Ans_list[Stu_num])):   # 迭代 Ans_list 裡的每個元素
            if(type(Ans_list[Stu_num][Que_num]) != int and len(Ans_list[Stu_num][Que_num]) > 1):
                # 如果收到的答案不只一個字母，就把所有答案合成連續的字母
                # eg. 收到'A, B, C' 轉成'ABC'

                temp_str = ""

                for n in range(0, len(Ans_list[Stu_num][Que_num]), 3):
                    temp_str = temp_str + str(Ans_list[Stu_num][Que_num][n])

                Ans_list[Stu_num][Que_num] = temp_str

        Stu_num += 1

    except Exception as e:
        # DEBUG
        print(str(e))

print("老師解答：")
print(Ans_list[0])
print("-----")

for k in range(1, stu_number+1):
    if(Ans_list[k][0] == None):  # 讀到底啦
        break

    stu_score = total_score      # 扣分制
    print("第", k, "位作答者")   # 流水號
    print(Ans_list[k])           # 他的答案

    # 確認兩次號碼的輸入都相同
    num1 = ws.cell(k+2, 3).value
    num2 = ws.cell(k+2, 4).value
    if(num1 != num2):
        print("填到不一樣的座號("+num1+" "+num2+")啦")
        print("-----")
        continue

    # mutiple chioce
    for i in range(0, num_choice):
        if(Ans_list[0][i] != Ans_list[k][i]):
            stu_score -= scr_choice

    # mutiple selection
    for i in range(num_choice, num_choice+num_selection):
        solution = mutiple_selection(
            Ans_list[0][i], Ans_list[k][i])
        if(solution == 1):
            stu_score -= scr_selection*2/5
        elif(solution == 2):
            stu_score -= scr_selection*4/5
        elif(solution >= 3):
            stu_score -= scr_selection

    student_score = round(stu_score, 2)     # 四捨五入
    student_class = ws.cell(k+2, 2).value   # 讀取表單上的學生班級
    student_number = ws.cell(k+2, 3).value  # 讀取表單上的學生座號

    print(student_class+"-"+student_number+" 成績：", student_score)
    print("-----")
