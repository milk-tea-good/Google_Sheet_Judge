# 開始縫合 輸入目標題數配分等等
# 好像壞掉了= =""

# -*- coding: utf-8 -*-
# 有ㄌ上面這行好像就可以打中文註解

# remember to 'pip install openpyxl'
from openpyxl import load_workbook
from selenium.webdriver.support import expected_conditions

wb = load_workbook('Test003.xlsx')

# grab the active worksheet (first)
ws = wb.active

# 計算多選題錯的選項數的函式


def mutiple_selection(ans, stu):
    # initialization
    option = []
    for a in range(0, 5):
        option.append(1)
    counter = 0

    for i in range(len(ans)):
        temp1 = ans[i:i+1]
        if(60 <= ord(temp1)):
            option[ord(temp1)-65] *= (-1)

    for i in range(len(stu)):
        temp2 = stu[i:i+1]
        if(60 <= ord(temp2)):
            option[ord(temp2)-65] *= (-1)

    for i in range(0, 5):
        if(option[i] != 1):
            counter -= -1

    return counter
    # 回傳錯的選項數


Stu_num = 0     # 座號（要注意是從0開始編號）
Que_num = 0     # 題號（要注意是從0開始編號）

num_choice = int(input("單選題總題數："))       # 單選題數

num_selection = int(input("多選題總題數："))    # 多選題數
print("單選：1～%d  多選：%d～%d " %
      (num_choice, num_choice+1, num_choice+num_selection))

total_score = int(input("此試券總分："))        # 總分

scr_choice = int(input("單選單題配分："))       # 單選配分

scr_selection = int(input("多選單題配分："))    # 多選配分
# compare with row[0]

stu_number = int(input("學生人數："))


# 這是最後要給沛儒的資料，二維陣列
Ans_list = [[0]*(num_choice+num_selection) for i in range(stu_number)]

# 45列(座號)、30行(題數)
# 引用方法： Ans_list[Stu_num][Que_num]

# for row in ws.iter_rows(min_row=, min_col=2, max_row=1+stu_number, max_col=1+num_choice+num_selection):

for X in range(2, 2+stu_number):
    for Y in range(5, 5+num_choice+num_selection):
        Ans_list[X-2][Y-5] = ws.cell(X, Y).value
        # 讀入.xlsx的答案的範圍

    for Que_num in range(len(Ans_list[X-2]) - 1):

        if(type(Ans_list[X-2][Que_num]) != int and len(Ans_list[X-2][Que_num]) > 1):
            # 如果收到的答案不只一個字母，就把所有答案合成連續的字母
            # eg. 收到'A, B, C' 轉成'ABC'

            temp_str = ""

            for n in range(0, len(Ans_list[X-2][Que_num]), 3):
                temp_str = temp_str + str(Ans_list[X-2][Que_num][n])

                Ans_list[X-2][Que_num] = temp_str
                # print(temp_str)

            # else:
            # print(Ans_list[Stu_num][Que_num])

    print(Ans_list[X-2])
