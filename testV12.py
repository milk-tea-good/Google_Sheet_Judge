
# 完整程式 with fill_in questions

from openpyxl import load_workbook

print('''
|-------------------------------|
| 記得將檔案名稱修改為 Test004  |
|-------------------------------|
''')

num_choice = int(input("單選題總題數："))       # 單選題數
scr_choice = int(input("單選單題配分："))       # 單選配分

num_selection = int(input("多選題總題數："))    # 多選題數
scr_selection = int(input("多選單題配分："))    # 多選配分

num_fillin = int(input("填充題總題數："))       # 填充題數
scr_fillin = int(input("填充單題配分："))       # 填充配分

print("單選：1～%d  多選：%d～%d  填充：%d～%d" %
      (num_choice, num_choice+1, num_choice+num_selection,
       num_choice+num_selection+1, num_choice+num_selection+num_fillin))

# compare with row[0]
stu_number = int(input("學生人數："))

wb = load_workbook('Test004.xlsx')      # 讀入 Test004.xlsx
ws = wb.active  # grab the active worksheet (first)

Ans_list = [[]]     # 宣告一個 list 以存放解答

# 讀入答案
m = 0
for i in range(1, stu_number+2):
    if (i != 1):
        Ans_list.append([])

    for j in range(9, 9+num_choice+num_selection):  # for 單選&多選
        temp = ws.cell(row=i, column=j).value
        temp = str(temp)                # 把 int -> str 統一化
        Ans_list[m].append(temp)

    for j in range(9+num_choice+num_selection, 9+num_choice+num_selection+num_fillin):
        temp = ws.cell(row=i, column=j).value   # 暫存資料做修正
        temp = str(temp)
        temp = temp.replace(' ', '')    # 消去空白格
        Ans_list[m].append(temp)        # 輸入到最後一項
    m += 1

# 選項結合
for i in range(0, len(Ans_list)):       # 對所有學生

    Que_num = 0
    for Que_num in range(num_choice, num_choice+num_selection):   # 迭代每個多選題
        if(type(Ans_list[i][Que_num]) != int and
           len(Ans_list[i][Que_num]) > 1):
            # 如果收到的答案不只一個字母，就把所有答案合成連續的字母
            # eg. 收到'A, B, C' 轉成'ABC'
            temp_str = ""
            temp_str = Ans_list[i][Que_num].replace(' ', '')    # 消去空白格
            temp_str = temp_str.replace(',', '')    # 消去逗號

            Ans_list[i][Que_num] = temp_str

# 全形轉成半形
for i in range(0, stu_number+1):
    for j in range(num_choice+num_selection, len(Ans_list[i])):
        for k in range(0, len(Ans_list[i][j])):
            if(ord(Ans_list[i][j][k]) > 65000):
                word = ord(Ans_list[i][j][k])-65248     # 全形轉回半形
                Ans_list[i][j] = Ans_list[i][j][:k] + \
                    chr(word) + Ans_list[i][j][k+1:]    # 組合字串


def mutiple_selection(ans, stu):    # 計算多選題錯的選項數的函式
    # initialization
    option = []
    for a in range(0, 5):
        option.append(1)
    counter = 0

    for i in range(len(ans)):
        temp1 = ans[i:i+1]
        if(60 <= ord(temp1)):
            option[ord(temp1)-65] *= (-1)

    for i in range(len(stu)):
        temp2 = stu[i:i+1]
        if(60 <= ord(temp2)):
            option[ord(temp2)-65] *= (-1)

    for i in range(0, 5):
        if(option[i] != 1):
            counter -= -1

    return counter      # 回傳錯的選項數


for k in range(1, stu_number+1):
    stu_score = 0

    # mutiple chioce
    for i in range(0, num_choice):
        if(Ans_list[0][i] == Ans_list[k][i]):
            stu_score += scr_choice

    # mutiple selection
    for i in range(num_choice, num_choice+num_selection):
        solution = mutiple_selection(Ans_list[0][i], Ans_list[k][i])
        if(solution == 0):
            stu_score += scr_selection
        elif(solution == 1):
            stu_score += scr_selection*0.6
        elif(solution == 2):
            stu_score += scr_selection*0.2

    # fill_in
    for i in range(num_choice+num_selection, num_choice+num_selection+num_fillin):
        if(Ans_list[0][i] == Ans_list[k][i]):
            stu_score += scr_fillin
            if(i == 8 or i == 9 or i == 10):  # 分數修正
                stu_score += 1

    student_score = round(stu_score, 2)     # 四捨五入
    ws.cell(k+1, 3, student_score)
    stu_name = ws.cell(k+1, 8).value
    print(stu_name, student_score)

wb.save('answer_file.xlsx')
print("已成功將資料輸出至 answer_file.xlsx ")
print()
print("按任意鍵以結束......")
input()  # 等待按鍵
